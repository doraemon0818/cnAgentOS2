"""
报表生成模块
支持生成 PDF 和 Excel 格式的报表
"""
import os
import json
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 报表存储目录
REPORT_DIR = os.path.join(os.path.dirname(__file__), '..', 'reports')
os.makedirs(REPORT_DIR, exist_ok=True)

# 注册中文字体
# Windows 系统字体路径
CHINESE_FONT = None

# 优先尝试注册黑体（TTF 格式，最简单）
try:
    simhei_path = r'C:\Windows\Fonts\simhei.ttf'
    if os.path.exists(simhei_path):
        pdfmetrics.registerFont(TTFont('SimHei', simhei_path))
        CHINESE_FONT = 'SimHei'
        print(f"成功注册中文字体: SimHei")
except Exception as e:
    print(f"注册 SimHei 失败: {e}")

# 如果黑体失败，尝试宋体（TTC 格式需要指定索引）
if not CHINESE_FONT:
    try:
        simsun_path = r'C:\Windows\Fonts\simsun.ttc'
        if os.path.exists(simsun_path):
            # TTC 文件包含多个字体，索引 0 是宋体
            pdfmetrics.registerFont(TTFont('SimSun', simsun_path, subfontIndex=0))
            CHINESE_FONT = 'SimSun'
            print(f"成功注册中文字体: SimSun")
    except Exception as e:
        print(f"注册 SimSun 失败: {e}")

# 如果宋体也失败，尝试微软雅黑
if not CHINESE_FONT:
    try:
        msyh_path = r'C:\Windows\Fonts\msyh.ttc'
        if os.path.exists(msyh_path):
            pdfmetrics.registerFont(TTFont('MicrosoftYaHei', msyh_path, subfontIndex=0))
            CHINESE_FONT = 'MicrosoftYaHei'
            print(f"成功注册中文字体: MicrosoftYaHei")
    except Exception as e:
        print(f"注册 MicrosoftYaHei 失败: {e}")

if not CHINESE_FONT:
    print("警告: 未找到可用的中文字体，PDF 报表中的中文可能显示为乱码")


class ReportGenerator:
    """报表生成器"""
    
    @staticmethod
    def generate_pdf_report(report_type: str, data: dict, name_prefix: str = None) -> str:
        """生成 PDF 报表"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if name_prefix:
            filename = f"{name_prefix}_{timestamp}.pdf"
        else:
            filename = f"{report_type}_{timestamp}.pdf"
        filepath = os.path.join(REPORT_DIR, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # 设置中文字体
        font_name = CHINESE_FONT or 'Helvetica'
        
        # 标题
        title_style = ParagraphStyle(
            'CustomTitle',
            fontSize=18,
            bold=True,
            alignment=1,
            spaceAfter=20,
            fontName=font_name
        )
        title_text = data.get('title', '报表')
        title = Paragraph(title_text, title_style)
        elements.append(title)
        
        # 副标题（日期）
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            fontSize=10,
            alignment=1,
            textColor=colors.gray,
            spaceAfter=20,
            fontName=font_name
        )
        subtitle_text = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        subtitle = Paragraph(subtitle_text, subtitle_style)
        elements.append(subtitle)
        
        elements.append(Spacer(1, 0.5*inch))
        
        # 添加统计卡片
        if 'stats' in data:
            stats = data['stats']
            stats_data = [
                ['统计项', '数值'],
            ]
            for key, value in stats.items():
                stats_data.append([key, str(value)])
            
            stats_table = Table(stats_data, colWidths=[2*inch, 3*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(stats_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # 添加数据表格
        if 'table_data' in data:
            table_data = data['table_data']
            table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            heading_style = ParagraphStyle(
                'CustomHeading2',
                fontSize=14,
                bold=True,
                spaceAfter=10,
                fontName=font_name
            )
            elements.append(Paragraph("详细数据", heading_style))
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
        
        # 添加图表数据（如果有）
        if 'chart_data' in data:
            heading_style = ParagraphStyle(
                'CustomHeading2',
                fontSize=14,
                bold=True,
                spaceAfter=10,
                fontName=font_name
            )
            elements.append(Paragraph("图表数据", heading_style))
            body_style = ParagraphStyle(
                'CustomBodyText',
                fontSize=10,
                fontName=font_name
            )
            chart_info = "\n".join([f"- {k}: {v}" for k, v in data['chart_data'].items()])
            elements.append(Paragraph(chart_info, body_style))
        
        doc.build(elements)
        return filepath
    
    @staticmethod
    def generate_excel_report(report_type: str, data: dict, name_prefix: str = None) -> str:
        """生成 Excel 报表"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if name_prefix:
            filename = f"{name_prefix}_{timestamp}.xlsx"
        else:
            filename = f"{report_type}_{timestamp}.xlsx"
        filepath = os.path.join(REPORT_DIR, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "报表数据"
        
        # 标题样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_text_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # 标题
        ws.cell(row=row, column=1, value=data.get('title', '报表')).font = title_font
        row += 1
        ws.cell(row=row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        row += 2
        
        # 添加统计卡片
        if 'stats' in data:
            stats = data['stats']
            ws.cell(row=row, column=1, value="统计信息").font = header_font
            row += 1
            
            ws.cell(row=row, column=1, value="统计项").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=2, value="数值").font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            row += 1
            
            for key, value in stats.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 1
        
        # 添加数据表格
        if 'table_data' in data:
            table_data = data['table_data']
            ws.cell(row=row, column=1, value="详细数据").font = header_font
            row += 1
            
            for r_idx, row_data in enumerate(table_data):
                for c_idx, cell_value in enumerate(row_data):
                    cell = ws.cell(row=row + r_idx, column=c_idx + 1, value=cell_value)
                    if r_idx == 0:
                        cell.font = header_font
                        cell.fill = header_fill
                    cell.border = thin_border
            
            row += len(table_data) + 1
        
        # 添加图表数据
        if 'chart_data' in data:
            ws.cell(row=row, column=1, value="图表数据").font = header_font
            row += 1
            for key, value in data['chart_data'].items():
                ws.cell(row=row, column=1, value=f"- {key}: {value}")
                row += 1
        
        # 调整列宽
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        wb.save(filepath)
        return filepath
    
    @staticmethod
    def generate_weekly_report(data: dict) -> str:
        """生成周报"""
        report_data = {
            'title': '周报',
            'stats': data.get('stats', {}),
            'table_data': data.get('table_data', []),
            'chart_data': data.get('chart_data', {})
        }
        return ReportGenerator.generate_pdf_report('weekly', report_data)
    
    @staticmethod
    def generate_daily_report(data: dict) -> str:
        """生成日报"""
        report_data = {
            'title': '日报',
            'stats': data.get('stats', {}),
            'table_data': data.get('table_data', []),
            'chart_data': data.get('chart_data', {})
        }
        return ReportGenerator.generate_pdf_report('daily', report_data)
    
    @staticmethod
    def get_report_list() -> list:
        """获取报表列表"""
        reports = []
        if not os.path.exists(REPORT_DIR):
            return reports
        
        for filename in os.listdir(REPORT_DIR):
            filepath = os.path.join(REPORT_DIR, filename)
            if os.path.isfile(filepath):
                stat = os.stat(filepath)
                reports.append({
                    'filename': filename,
                    'filepath': filepath,
                    'size': stat.st_size,
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                })
        
        # 按创建时间倒序排序
        reports.sort(key=lambda x: x['created_time'], reverse=True)
        return reports
    
    @staticmethod
    def get_report_path(filename: str) -> str:
        """获取报表文件路径"""
        return os.path.join(REPORT_DIR, filename)
    
    @staticmethod
    def generate_task_report(task_id: int, task_name: str, task_type: str) -> dict:
        """根据任务生成报表数据"""
        report_data = {
            'title': f'{task_name} - 任务报表',
            'stats': {
                '任务ID': str(task_id),
                '任务名称': task_name,
                '任务类型': task_type,
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            },
            'table_data': [
                ['项目', '状态', '数据量'],
                ['数据采集', '成功', '150'],
                ['数据处理', '成功', '120'],
                ['报表生成', '成功', '1'],
            ],
            'chart_data': {
                '总数据量': 150,
                '有效数据': 120,
                '成功率': '80%'
            }
        }
        return report_data
